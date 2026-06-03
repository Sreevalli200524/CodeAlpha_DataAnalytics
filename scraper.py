import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

base_url = "https://books.toscrape.com/catalogue/page-{}.html"

all_books = []

# Scrape first 5 pages
for page in range(1, 6):

    print(f"Scraping Page {page}...")

    url = base_url.format(page)

    response = requests.get(url)

    soup = BeautifulSoup(response.text, "html.parser")

    books = soup.find_all("article", class_="product_pod")

    for book in books:

        # Title
        title = book.h3.a["title"]

        # Price
        price = book.find("p", class_="price_color").text
        price = price.replace("Â", "").replace("£", "")
        price = float(price)

        # Rating Conversion
        rating_text = book.p["class"][1]

        rating_map = {
            "One": 1,
            "Two": 2,
            "Three": 3,
            "Four": 4,
            "Five": 5
        }

        rating = rating_map.get(rating_text, 0)

        # Availability
        availability = book.find(
            "p",
            class_="instock availability"
        ).text.strip()

        if "In stock" in availability:
            stock = "Available"
        else:
            stock = "Out of Stock"

        # Price Category
        if price < 20:
            category = "Budget"
        elif price < 35:
            category = "Standard"
        else:
            category = "Premium"

        # Store Data
        all_books.append({
            "Book Title": title,
            "Price (£)": round(price, 2),
            "Price Category": category,
            "Rating": rating,
            "Stock Status": stock,
            "Page Number": page
        })

    time.sleep(1)

# Create DataFrame
df = pd.DataFrame(all_books)

# Sort by price
df = df.sort_values(by="Price (£)", ascending=False)

# Save Excel File
file_name = "Professional_Books_Dataset.xlsx"

df.to_excel(file_name, index=False)

# Open workbook
wb = load_workbook(file_name)
ws = wb.active

# Header styling
header_fill = PatternFill(
    start_color="FF6B35",
    end_color="FF6B35",
    fill_type="solid"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

# Apply header styles
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Auto column width
for column in ws.columns:

    max_length = 0
    column_letter = get_column_letter(column[0].column)

    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass

    adjusted_width = max_length + 4
    ws.column_dimensions[column_letter].width = adjusted_width

# Freeze top row
ws.freeze_panes = "A2"

# Add filters
ws.auto_filter.ref = ws.dimensions

# Save workbook
wb.save(file_name)

print("\nPROFESSIONAL DATASET CREATED!")
print(f"File Saved As: {file_name}")
print(f"Total Records: {len(df)}")-----